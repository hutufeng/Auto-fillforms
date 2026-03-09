# -*- coding: utf-8 -*-
"""模板填充模块：支持 Word (docxtpl) 和 Excel (openpyxl) 模板"""

import os
import re
from docxtpl import DocxTemplate
from openpyxl import load_workbook


# 占位符 → 数据字典 key 的映射
PLACEHOLDER_MAP = {
    "姓名": "name",
    "性别": "gender",
    "民族": "ethnicity",
    "出生日期": "birthday",
    "年龄": "age",
    "身份证号码": "id_number",
    "住址": "address",
    "签发机关": "authority",
    "有效期限": "validity",
}


def fill_word_template(template_path: str, data: dict, output_path: str):
    """
    使用 docxtpl 填充 Word 模板。
    模板中使用 {{姓名}}、{{身份证号码}} 等 Jinja2 占位符。
    """
    doc = DocxTemplate(template_path)

    # 构建模板上下文：占位符名称 → 值
    context = {}
    for placeholder, key in PLACEHOLDER_MAP.items():
        value = data.get(key, "")
        # age 转为字符串
        context[placeholder] = str(value) if isinstance(value, int) else value

    doc.render(context)
    doc.save(output_path)


def fill_excel_template(template_path: str, data: dict, output_path: str):
    """
    填充 Excel 模板。
    模板中使用 {{姓名}}、{{身份证号码}} 等占位符，
    遍历所有单元格查找并替换。
    """
    wb = load_workbook(template_path)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    new_value = cell.value
                    for placeholder, key in PLACEHOLDER_MAP.items():
                        tag = "{{" + placeholder + "}}"
                        if tag in new_value:
                            value = data.get(key, "")
                            replacement = str(value) if isinstance(value, int) else value
                            new_value = new_value.replace(tag, replacement)
                    if new_value != cell.value:
                        cell.value = new_value

    wb.save(output_path)


def fill_template(template_path: str, data: dict, output_path: str):
    """
    根据模板文件扩展名自动选择填充方式。
    """
    ext = os.path.splitext(template_path)[1].lower()
    if ext == ".docx":
        fill_word_template(template_path, data, output_path)
    elif ext in (".xlsx", ".xls"):
        fill_excel_template(template_path, data, output_path)
    else:
        raise ValueError(f"不支持的模板格式: {ext}")


def batch_fill_template(template_path: str, data_list: list[dict], output_dir: str) -> list[str]:
    """
    批量填充模板，每人生成一份独立文件。
    
    Returns:
        生成的文件路径列表
    """
    ext = os.path.splitext(template_path)[1].lower()
    os.makedirs(output_dir, exist_ok=True)

    output_files = []
    for data in data_list:
        name = data.get("name", "未知")
        filename = f"{name}_filled{ext}"
        output_path = os.path.join(output_dir, filename)

        # 避免文件名重复
        counter = 1
        while os.path.exists(output_path):
            filename = f"{name}_filled_{counter}{ext}"
            output_path = os.path.join(output_dir, filename)
            counter += 1

        fill_template(template_path, data, output_path)
        output_files.append(output_path)

    return output_files
