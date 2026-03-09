# -*- coding: utf-8 -*-
"""Excel 数据存储模块：支持创建和追加"""

import os
from openpyxl import Workbook, load_workbook

# 表头字段定义
HEADERS = ["姓名", "性别", "民族", "出生日期", "年龄", "身份证号码", "住址", "签发机关", "有效期限"]

# 数据字典 key → 表头对应顺序
FIELD_MAP = ["name", "gender", "ethnicity", "birthday", "age", "id_number", "address", "authority", "validity"]


def _data_to_row(data: dict) -> list:
    """将数据字典转为与 HEADERS 对应的行列表"""
    return [data.get(key, "") for key in FIELD_MAP]


def write_to_excel(data: dict, output_path: str):
    """
    将单条身份证数据写入 Excel。
    文件不存在则创建，已存在则追加到行尾。
    """
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "身份证信息"
        ws.append(HEADERS)
        # 设置列宽
        col_widths = [10, 6, 8, 14, 6, 22, 40, 20, 26]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.append(_data_to_row(data))
    wb.save(output_path)


def write_batch_to_excel(data_list: list[dict], output_path: str):
    """
    批量写入多条身份证数据到 Excel。
    """
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "身份证信息"
        ws.append(HEADERS)
        col_widths = [10, 6, 8, 14, 6, 22, 40, 20, 26]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for data in data_list:
        ws.append(_data_to_row(data))
    wb.save(output_path)
